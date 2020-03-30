import os
import sys
import mouse
import logging
import keyboard
import datetime
import openpyxl
import win32gui
import threading
import resource_rc
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QSystemTrayIcon, QMenu


logging.basicConfig(level=logging.DEBUG, format='(%(threadName)-9s) %(message)s',)


class TimeStamp(threading.Thread):
    delta = 3
    prev  = datetime.datetime(2020, 1, 1, 0, 0, 0)

    def __init__(self):
        threading.Thread.__init__(self)
        self.setDaemon(True)

    
    def run(self):
        #logging.info(str('start'))
        os.chdir(os.path.expanduser("~/Desktop"))
        if not os.path.exists("WORK_TIME"):
            os.mkdir("WORK_TIME")
        os.chdir("./WORK_TIME")

        mouse.hook(self.TimeCheck)
        keyboard.hook(self.TimeCheck)
    
        win32gui.PumpMessages()


    def TimeCheck(self, key):
        #logging.info(str('time checking..'))
        now = datetime.datetime.now()

        interval = now - self.prev

        if interval.seconds >= self.delta:
            self.prev = now
            self.WriteToExcel(now)


    def WriteToExcel(self, now):
        #logging.info(str('writing..'))
        nYear  = str(now.year)
        nMonth = str(now.month)
        nDay   = str(now.day)
        nHour  = str(now.hour)
        nMin   = str(now.minute)
        nSec   = str(now.second)

        fileName = nYear + "_" + nMonth + ".xlsx"
        if not os.path.isfile(fileName):
            xlFile = self.MakeTemplate(fileName)
        else:
            xlFile = openpyxl.load_workbook(fileName)

        sheet = xlFile['Sheet']

        if not sheet.cell(row = sheet.max_row, column = 1).value == nDay + "일":
            sheet.cell(row = sheet.max_row + 1, column = 1).value = nDay + "일"
            sheet.cell(row = sheet.max_row, column = 2).value = nHour + ":" + nMin + ":" + nSec
            sheet.cell(row = sheet.max_row, column = 4).number_format = "hh.mm.ss"
            sheet.cell(row = sheet.max_row, column = 4).value = "=C" + str(sheet.max_row) + "-B" + str(sheet.max_row)

        sheet.cell(row = sheet.max_row, column = 3).value = nHour + ":" + nMin + ":" + nSec

        xlFile.save(fileName)
        xlFile.close()


    def MakeTemplate(self, fileName):
        xlFile = openpyxl.Workbook()

        sheet = xlFile['Sheet']

        sheet.cell(row = 1, column = 1).value = 'day'
        sheet.cell(row = 1, column = 2).value = 'start time'
        sheet.cell(row = 1, column = 3).value = 'end time'
        sheet.cell(row = 1, column = 4).value = 'working time'

        xlFile.save(fileName)

        return xlFile


class SystemTrayIcon(QSystemTrayIcon):
    def __init__(self, icon, parent):
        QSystemTrayIcon.__init__(self, icon, parent)

        self.setToolTip('Check out my tray icon')

        menu = QMenu()
        exitAction = menu.addAction('Exit')
        exitAction.triggered.connect(parent.quit)

        TimeStamp().start()

        self.setContextMenu(menu)


def main():
    app = QApplication(sys.argv)

    trayIcon = SystemTrayIcon(QIcon(':image/mainIcon.png'), parent = app)
    trayIcon.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()